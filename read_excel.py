from openpyxl import load_workbook

wb = load_workbook(r'C:\Users\shiueo\Downloads\weawefawef.xlsx')
ws = wb.active

for row in ws.iter_rows(values_only=True):
    print(row)
